from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
import tweet_util




def main():
	tweets= []
	wb = load_workbook('MorgansTweetClassifications.xlsx')
	print(wb.get_sheet_names())
	#ws = wb.worksheet[0]
	ws = wb.get_sheet_by_name('Sheet1')
	#for col_idx in xrange(2,4):
	#	col = get_column_letter(col_idx)
	for row in xrange(1,376):
		tmp = []
		tmp.append(ws.cell('%s%s'%(get_column_letter(1),row)).value)
		tmp.append(ws.cell('%s%s'%(get_column_letter(2),row)).value)
		tmp.append(ws.cell('%s%s'%(get_column_letter(3),row)).value)
		#print(tmp)
		tweets.append(tmp)
		#print(ws.cell('%s%s'%(get_column_letter(2),row)).value+' | '+ws.cell('%s%s'%(get_column_letter(3),row)).value)
	print(tweets)

if __name__ == "__main__":
	main()
